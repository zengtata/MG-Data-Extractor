import os
from datetime import datetime  # Import to work with dates and times
from tkinter import messagebox  # Import to show message boxes (pop-ups)

import openpyxl  # Library for working with Excel files
from openpyxl import Workbook  # Import Workbook to create new Excel workbooks
from openpyxl.styles import PatternFill, Font, Border, Side  # Import styles for formatting Excel cells

def process_cipl(input_files, output_filename):
    # This function processes multiple Excel files (input_files) and extracts specific data to save into an output file (output_filename).

    def extract_data(sheet1):
        # This nested function extracts data from the 'CI' sheet of an Excel workbook.

        # Initialize variables to store extracted data
        invoiceNO, date_value, sale_contractNO, seller, to_value, delivery_term, total_unit, total_value, delivery_number = "", "", "", "", "", "", "", "", ""

        # Iterate over each row in the sheet, with values_only=True to get cell values directly
        for row in sheet1.iter_rows(values_only=True):
            for idx, cell in enumerate(row):  # Iterate over each cell in the row
                # Check for specific labels in the cells and extract data accordingly
                if cell == "INVOICE NO.:":
                    invoiceNO = row[idx + 1]  # Get the next cell value for the invoice number
                elif cell == "DATE:":
                    date_value = row[idx + 1]
                    if isinstance(date_value, datetime):
                        date_value = date_value.strftime('%Y-%m-%d')  # Convert the date to a string in YYYY-MM-DD format
                elif cell == "SALE CONTRACT NO.:":
                    sale_contractNO = row[idx + 1]  # Get the sale contract number
                elif cell == "SELLER: ":
                    # Extract the seller information by combining multiple cells until a None value is encountered
                    seller = []
                    for next_idx in range(idx + 1, len(row)):
                        if row[next_idx] is None:
                            break
                        seller.append(row[next_idx])
                    seller = " ".join(seller)  # Join the parts of the seller information into a single string
                elif cell == "TO:":
                    # Extract the 'TO' value similarly by combining multiple cells
                    to_values = []
                    for next_idx in range(idx + 1, len(row)):
                        if row[next_idx] is None:
                            break
                        to_values.append(row[next_idx])
                    to_value = " ".join(to_values)
                elif cell == "DELIVERY TERM:":
                    delivery_term = row[idx + 1]  # Get the delivery term
                elif cell == "TOTAL":
                    total_unit = row[idx + 3]  # Get the total unit value, which is 3 cells after "TOTAL"
                elif cell is None and idx + 2 < len(row) and row[idx + 1] == "EUR" and row[idx - 1] is None:
                    total_value = row[idx + 2]  # Extract the total value if specific conditions are met
                elif isinstance(cell, str) and cell.startswith("DELIVERY NO.:"):
                    # Extract the delivery number if the cell starts with "DELIVERY NO.:"
                    parts = cell.split(":")
                    if len(parts) > 1:
                        delivery_number = parts[1].strip()  # Get the part after the colon and strip any whitespace

        # Return a dictionary with all the extracted values
        return {
            "Invoice_NO": invoiceNO,
            "Date": date_value,
            "Sale_contract_NO": sale_contractNO,
            "Seller": seller,
            "To": to_value,
            "Delivery_term": delivery_term,
            "Invoice_total_unit": total_unit,
            "Invoice_total_value": total_value,
            "Delivery_number": delivery_number
        }

    def extract_vin_numbers(sheet2):
        # This nested function extracts VIN numbers from the 'PL' sheet of an Excel workbook.
        vin_numbers = set()  # Use a set to store VIN numbers and avoid duplicates

        # Iterate over each column in the sheet
        for col in sheet2.iter_cols(values_only=True):
            for idx, cell in enumerate(col):  # Iterate over each cell in the column
                if isinstance(cell, str) and cell.startswith("LS") and len(cell) == 17:
                    vin_numbers.add(cell)  # Add VIN numbers that match the specific format to the set

        vin_numbers = list(vin_numbers)  # Convert the set back to a list for further processing

        return vin_numbers

    # Check if the output file already exists
    if os.path.exists(output_filename):
        output_workbook = openpyxl.load_workbook(output_filename)  # Load the existing workbook
        if "CIPL_extracted_data" in output_workbook.sheetnames:
            new_sheet = output_workbook["CIPL_extracted_data"]  # Use the existing sheet if it already exists
        else:
            # Create a new sheet and add header titles with formatting
            new_sheet = output_workbook.create_sheet(title="CIPL_extracted_data")
            titles = ["VIN Numbers", "Invoice_NO", "Date", "Sale_contract_NO", "Seller", "To", "Delivery_term",
                      "Invoice_total_unit", "Invoice_total_value", "Delivery_number"]
            for col_idx, title in enumerate(titles, start=1):
                cell = new_sheet.cell(row=1, column=col_idx, value=title)
                # Apply a light green fill to the header cells
                cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                cell.font = Font(bold=True)  # Make the header text bold
                # Add thin borders around the header cells
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                                     bottom=Side(style='thin'))
    else:
        # If the output file doesn't exist, create a new workbook and sheet
        output_workbook = Workbook()
        output_workbook.remove(output_workbook.active)  # Remove the default sheet
        new_sheet = output_workbook.create_sheet(title="CIPL_extracted_data")
        titles = ["VIN Numbers", "Invoice_NO", "Date", "Sale_contract_NO", "Seller", "To", "Delivery_term",
                  "Invoice_total_unit", "Invoice_total_value", "Delivery_number"]
        for col_idx, title in enumerate(titles, start=1):
            cell = new_sheet.cell(row=1, column=col_idx, value=title)
            cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Apply the same formatting
            cell.font = Font(bold=True)
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                                 bottom=Side(style='thin'))

    row_idx = new_sheet.max_row + 1  # Start adding data from the next available row

    # Loop through each input file to process and extract the data
    for input_file in input_files:
        try:
            workbook = openpyxl.load_workbook(input_file, data_only=True)  # Load the input workbook
            sheet1 = workbook['CI']  # Access the 'CI' sheet
            data = extract_data(sheet1)  # Extract data from 'CI' sheet
            sheet2 = workbook['PL']  # Access the 'PL' sheet
            vin_numbers = extract_vin_numbers(sheet2)  # Extract VIN numbers from 'PL' sheet

            for vin in vin_numbers:
                data_exists = False
                # Check if the data already exists in the output sheet to avoid duplicates
                for row in new_sheet.iter_rows(min_row=2, max_row=new_sheet.max_row, values_only=True):
                    if row[0] == vin and row[1] == data["Invoice_NO"] and row[2] == data["Date"] and \
                            row[3] == data["Sale_contract_NO"] and row[4] == data["Seller"] and \
                            row[5] == data["To"] and row[6] == data["Delivery_term"] and \
                            row[7] == data["Invoice_total_unit"] and row[8] == data["Invoice_total_value"] and \
                            row[9] == data["Delivery_number"]:
                        data_exists = True
                        break

                if not data_exists:
                    new_sheet.cell(row=row_idx, column=1, value=vin)  # Write the VIN number in the first column
                    col_idx = 2  # Start filling from the second column
                    for key in ["Invoice_NO", "Date", "Sale_contract_NO", "Seller", "To", "Delivery_term",
                                "Invoice_total_unit", "Invoice_total_value", "Delivery_number"]:
                        cell = new_sheet.cell(row=row_idx, column=col_idx, value=data[key])  # Write each piece of extracted data
                        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                                             bottom=Side(style='thin'))  # Apply borders to each cell
                        col_idx += 1
                    row_idx += 1  # Move to the next row

        except Exception as e:
            # Handle any errors that occur during processing
            print(f"Failed to process {input_file}: {e}")
            messagebox.showwarning("File Error", f"Failed to process {input_file}: {e}")  # Show a warning message

    # Apply thin borders to all cells in the sheet for consistent formatting
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                         bottom=Side(style='thin'))

    for row in new_sheet.iter_rows():
        for cell in row:
            cell.border = thin_border

    output_workbook.save(output_filename)  # Save the output workbook
    messagebox.showinfo("Process Complete",
                        f"Data has been saved to {output_filename}\nNumber of rows created: {row_idx - 2}")  # Show an info message when done
