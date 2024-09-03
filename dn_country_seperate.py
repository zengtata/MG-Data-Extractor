import os  # Import os module to handle file and directory operations
from tkinter import messagebox  # Import messagebox to display messages to the user

import openpyxl  # Import openpyxl to work with Excel files
from openpyxl import Workbook  # Import Workbook to create new Excel workbooks

def process_dn_seperate(file_path, output_base_dir):
    # This function processes an Excel file, separates the data based on the 'Dest.' column,
    # and saves the data into separate files in a specified output directory.

    # Load the input workbook from the provided file path
    wb_input = openpyxl.load_workbook(file_path)
    # Select the sheet named 'Main' in the workbook
    sheet_input = wb_input['Main']

    # Read the headers from the first row of the sheet
    headers = [cell.value for cell in sheet_input[1]]
    # Create a dictionary to map header names to their column index
    header_to_index = {header: idx for idx, header in enumerate(headers)}

    # Check if the 'Dest.' column exists in the headers
    if 'Dest.' not in header_to_index:
        # If 'Dest.' column is not found, raise an error
        raise ValueError("The column 'Dest.' was not found in the input sheet.")

    # Get the index of the 'Dest.' column
    dest_column_index = header_to_index['Dest.']

    # Initialize a dictionary to collect rows of data for each destination
    data_by_dest = {}

    # Iterate over the rows of the sheet, starting from the second row to skip headers
    for row in sheet_input.iter_rows(min_row=2, values_only=True):
        dest = row[dest_column_index]  # Get the value from the 'Dest.' column for the current row
        if dest not in data_by_dest:
            data_by_dest[dest] = []  # Create a new list for this destination if it doesn't exist
        data_by_dest[dest].append(row)  # Append the current row to the list for this destination

    # Create a new directory for the output files
    base_name = os.path.splitext(os.path.basename(file_path))[0]  # Get the base name of the input file
    output_dir = os.path.join(output_base_dir, base_name)  # Combine the base name with the output directory path
    os.makedirs(output_dir, exist_ok=True)  # Create the directory if it doesn't exist

    # Iterate over each destination and its associated rows
    for dest, rows in data_by_dest.items():
        # Create a safe file name for the destination, replacing invalid characters
        dest_safe_name = "".join(c if c.isalnum() or c in (" ", "_") else "_" for c in dest)[:31]
        # Create the full output file path by combining the directory and file name
        output_file_path = os.path.join(output_dir, f"{base_name}_{dest_safe_name}.xlsx")

        # Create a new workbook for this destination's data
        wb_output = Workbook()
        sheet_output = wb_output.active  # Get the active sheet in the new workbook
        sheet_output.title = dest_safe_name  # Set the sheet title to the destination's safe name

        # Write the header row to the new sheet
        for col, header in enumerate(headers, start=1):
            sheet_output.cell(row=1, column=col, value=header)

        # Write the data rows to the new sheet
        for row_index, row_data in enumerate(rows, start=2):  # Start from row 2 to skip the header
            for col_index, cell_value in enumerate(row_data, start=1):
                sheet_output.cell(row=row_index, column=col_index, value=cell_value)

        # Save the new workbook to the specified file path
        wb_output.save(output_file_path)

    # Show a message box to the user indicating that the process is complete
    messagebox.showinfo("Success",
                        f"Data has been separated into new files based on the 'Dest.' column and saved in {output_dir}.")
