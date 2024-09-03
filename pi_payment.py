import os
from datetime import datetime
from tkinter import messagebox

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side


def process_pi_payment(input_files, output_filename):
    # Helper function to extract data from the sheet
    def extract_data(sheet):
        invoiceNO, date_value, importer, total_quantity, total_payment = "", "", "", "", ""
        tt, lc, tt_value, lc_value, currency, port_of_unload = "", "", "", "", "", ""
        material_code, model_quantity, unit_price, model = [], [], [], []

        for row in sheet.iter_rows(values_only=True):
            for idx, cell in enumerate(row):
                if isinstance(cell, str):
                    if cell == "Date:":
                        date_value = row[idx + 1]
                        if isinstance(date_value, datetime):
                            date_value = date_value.strftime('%Y-%m-%d')
                    elif cell.startswith("Invoice Number:"):
                        parts = cell.split(":")
                        if len(parts) > 1:
                            invoiceNO = parts[1].strip()
                    elif cell == "Port of Unloading:":
                        port_of_unload = row[idx + 2]
                    elif cell == "TOTAL Qty:":
                        for next_idx in range(idx + 1, len(row)):
                            if row[next_idx] is not None:
                                total_quantity = row[next_idx]
                                break
                    elif cell == "TOTAL PAYMENT :":
                        for next_idx in range(idx + 1, len(row)):
                            if row[next_idx] is not None:
                                total_payment = row[next_idx]
                                break
                    elif "T/T" in cell:
                        parts = cell.split(" ")
                        if len(parts) > 1:
                            tt = parts[0].strip()
                        for next_idx in range(idx + 1, len(row)):
                            if row[next_idx] is not None:
                                tt_value = row[next_idx]
                                break
                    elif "L/C" in cell:
                        parts = cell.split(" ")
                        if len(parts) > 1:
                            lc = parts[0].strip()
                        for next_idx in range(idx + 1, len(row)):
                            if row[next_idx] is not None:
                                lc_value = row[next_idx]
                                break
                    elif cell == "Currency：":
                        currency = row[idx + 1]

        for col in sheet.iter_cols(values_only=True):
            for idx, cell in enumerate(col):
                if isinstance(cell, str):
                    if cell == "Importer:":
                        for next_idx in range(idx + 1, len(col)):
                            if col[next_idx] is not None:
                                importer = col[next_idx]
                                break
                    elif cell == "Material Code":
                        for next_idx in range(idx + 1, len(col)):
                            if col[next_idx] is not None:
                                material_code.append(col[next_idx])
                            else:
                                break
                    elif cell == "Qty":
                        for next_idx in range(idx + 1, len(col)):
                            if col[next_idx] is not None:
                                model_quantity.append(col[next_idx])
                            else:
                                break
                    elif cell == "Unit Price":
                        for next_idx in range(idx + 1, len(col)):
                            if col[next_idx] is not None:
                                unit_price.append(col[next_idx])
                            else:
                                break

        return {
            "Invoice_NO.": invoiceNO,
            "Date": date_value,
            "Importer": importer,
            "Total Quantity": total_quantity,
            "Total Payment": total_payment,
            "T/T": tt,
            "L/C": lc,
            "T/T Value": tt_value,
            "L/C Value": lc_value,
            "Currency": currency,
            "Port of Unloading": port_of_unload,
            "Material Code": material_code,
            "Model Quantity": model_quantity,
            "Unit Price": unit_price,
        }

    # Check and create or load output workbook
    if os.path.exists(output_filename):
        output_workbook = openpyxl.load_workbook(output_filename)
        new_sheet = output_workbook[
            "PI_extracted_data"] if "PI_extracted_data" in output_workbook.sheetnames else output_workbook.create_sheet(
            title="PI_extracted_data")
    else:
        output_workbook = Workbook()
        output_workbook.remove(output_workbook.active)
        new_sheet = output_workbook.create_sheet(title="PI_extracted_data")
        titles = ["Invoice_NO.", "Date", "Importer", "Port of Unloading", "Total Quantity", "Total Payment",
                  "T/T", "T/T Value", "L/C", "L/C Value", "Currency"]
        for col_idx, title in enumerate(titles, start=1):
            cell = new_sheet.cell(row=1, column=col_idx, value=title)
            cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
            cell.font = Font(bold=True)
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                                 bottom=Side(style='thin'))

    detailed_sheet = output_workbook[
        "detailed_data"] if "detailed_data" in output_workbook.sheetnames else output_workbook.create_sheet(
        title="detailed_data")
    detailed_titles = ["Material Code", "Model Quantity", "Unit Price", "Invoice_NO."]
    for col_idx, title in enumerate(detailed_titles, start=1):
        cell = detailed_sheet.cell(row=1, column=col_idx, value=title)
        cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        cell.font = Font(bold=True)
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                             bottom=Side(style='thin'))

    existing_invoice_nos = {row[0] for row in new_sheet.iter_rows(min_row=2, values_only=True)}
    existing_detailed_data = {(row[0], row[1], row[2], row[3]) for row in
                              detailed_sheet.iter_rows(min_row=2, values_only=True)}

    row_idx = new_sheet.max_row + 1
    detailed_row_idx = detailed_sheet.max_row + 1

    for input_file in input_files:
        try:
            workbook = openpyxl.load_workbook(input_file, data_only=True)
            sheet = workbook.active
            data = extract_data(sheet)

            if data["Invoice_NO."] not in existing_invoice_nos:
                new_sheet.cell(row=row_idx, column=1, value=data["Invoice_NO."])
                new_sheet.cell(row=row_idx, column=2, value=data["Date"])
                new_sheet.cell(row=row_idx, column=3, value=data["Importer"])
                new_sheet.cell(row=row_idx, column=4, value=data["Port of Unloading"])
                new_sheet.cell(row=row_idx, column=5, value=data["Total Quantity"])
                new_sheet.cell(row=row_idx, column=6, value=data["Total Payment"])
                new_sheet.cell(row=row_idx, column=7, value=data["T/T"])
                new_sheet.cell(row=row_idx, column=8, value=data["T/T Value"])
                new_sheet.cell(row=row_idx, column=9, value=data["L/C"])
                new_sheet.cell(row=row_idx, column=10, value=data["L/C Value"])
                new_sheet.cell(row=row_idx, column=11, value=data["Currency"])
                existing_invoice_nos.add(data["Invoice_NO."])
                row_idx += 1

            for i in range(len(data["Material Code"])):
                if i < len(data["Model Quantity"]) and i < len(data["Unit Price"]):
                    detailed_data = (
                    data["Material Code"][i], data["Model Quantity"][i], data["Unit Price"][i], data["Invoice_NO."])
                    if detailed_data not in existing_detailed_data:
                        detailed_sheet.cell(row=detailed_row_idx, column=1, value=data["Material Code"][i])
                        detailed_sheet.cell(row=detailed_row_idx, column=2, value=data["Model Quantity"][i])
                        detailed_sheet.cell(row=detailed_row_idx, column=3, value=data["Unit Price"][i])
                        detailed_sheet.cell(row=detailed_row_idx, column=4, value=data["Invoice_NO."])
                        existing_detailed_data.add(detailed_data)
                        detailed_row_idx += 1

        except Exception as e:
            print(f"Failed to process {input_file}: {e}")
            messagebox.showwarning("File Error", f"Failed to process {input_file}: {e}")

    # Apply border to all cells
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                         bottom=Side(style='thin'))

    for sheet in [new_sheet, detailed_sheet]:
        for row in sheet.iter_rows():
            for cell in row:
                cell.border = thin_border

    output_workbook.save(output_filename)
    messagebox.showinfo("Process Complete",
                        f"Data has been saved to {output_filename}\n"
                        f"Number of rows created in main sheet: {row_idx - 2}\n"
                        f"Number of rows created in detailed sheet: {detailed_row_idx - 2}")

# The rest of the GUI code for file selection and execution remains the same
